import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

class ExcelHandler:
    def __init__(self):
        self.workbook = None
        self.sheet = None

    def generate_excel(self, list_cvs, output_path):
        """
        Gera um arquivo Excel estilizado com os dados dos currículos

        Args:
            curricula_list: Lista de objetos da classe Curriculum
            output_path: Local de envio do Excel
            output_file: Nome do arquivo Excel de saída
        """

        # Criar workbook e sheet
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Currículos"

        # Definir cabeçalhos
        headers = [
            "Nome", "Email", "Telefone",
            "Primeira Empresa", "Primeiro Cargo", "Primeiro Período",
            "Segunda Empresa", "Segundo Cargo", "Segundo Período"
        ]

        # Estilizar cabeçalhos
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Definir bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplicar cabeçalhos e estilo
        for col_idx, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Preencher dados dos currículos
        for row_idx, curriculum in enumerate(list_cvs, 2):
            # Dados básicos
            self.sheet.cell(row=row_idx, column=1).value = curriculum.nome
            self.sheet.cell(row=row_idx, column=2).value = curriculum.email
            self.sheet.cell(row=row_idx, column=3).value = curriculum.telefones

            # Primeira experiência
            self.sheet.cell(row=row_idx, column=4).value = curriculum.primeiraempresa
            self.sheet.cell(row=row_idx, column=5).value = curriculum.primeirocargo
            self.sheet.cell(row=row_idx, column=6).value = curriculum.primeiroperiodo

            # Segunda experiência
            self.sheet.cell(row=row_idx, column=7).value = curriculum.segundaempresa
            self.sheet.cell(row=row_idx, column=8).value = curriculum.segundocargo
            self.sheet.cell(row=row_idx, column=9).value = curriculum.segundoperiodo

            # Aplicar bordas e alinhamento às células de dados
            for col_idx in range(1, 10):
                cell = self.sheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Auto-dimensionar colunas para ajustar ao conteúdo
        for col_idx in range(1, 10):
            column_letter = get_column_letter(col_idx)
            # Calcular a largura baseada no conteúdo mais longo na coluna
            max_length = 0
            for row_idx in range(1, len(list_cvs) + 2):  # +2 para incluir o cabeçalho
                cell_value = self.sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Definir a largura da coluna (com um pouco de espaço extra)
            adjusted_width = max_length + 4
            self.sheet.column_dimensions[column_letter].width = adjusted_width

        # Congelar o cabeçalho
        self.sheet.freeze_panes = "A2"

        # Salvar o arquivo
        if not str(output_path).endswith(".xlsx"):
            output_path = output_path + ".xlsx"

        self.workbook.save(output_path)
        return
